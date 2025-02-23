import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnum).value

def writeData(file,sheetName,rowNum,columNum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rowNum,columNum).value=data
    workbook.save(file)

def fillGreenColor(file,sheetName,rowNum,columNum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rowNum,columNum).fill=greenFill
    workbook.save(file)

def fillRedColor(file,sheetName,rowNum,columNum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    redFill=PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rowNum,columNum).fill=redFill
    workbook.save(file)