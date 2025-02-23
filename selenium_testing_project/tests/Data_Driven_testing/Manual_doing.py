import openpyxl

file="D:\\Sai Work\\Python\\Practice\\sample-file.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"] ## OR sheet=workbook.active........ For current sheet
##sheet=workbook.create_sheet["Sheet3"]

rows=sheet.max_row
columns=sheet.max_column

#print(rows)

####### Reading data in a sheet ##########
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value, end="   ")

    print()    

####### Writing Data in sheet for same data ########
sheet=workbook["Sheet3"]
for r in range(1,4):
    for c in range(1,7):
        sheet.cell(r,c).value="Welcom!"
workbook.save(file)

####### Writing Data in sheet for Multiple data #######
sheet=workbook["Sheet4"]

sheet.cell(1,1).value="Name"
sheet.cell(1,2).value="Age"
sheet.cell(1,3).value="Job"
sheet.cell(1,4).value="Salary"

sheet.cell(2,1).value="Sai Krishna"
sheet.cell(2,2).value=28
sheet.cell(2,3).value="Software"
sheet.cell(2,4).value=69580.67

workbook.save(file)

print(sheet.max_row)
print(sheet.max_column)
print(rows)